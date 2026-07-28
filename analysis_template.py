#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""独立经营分析模板生成器及其 Tkinter 窗口。"""

from __future__ import annotations

import queue
import re
import threading
import tkinter as tk
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

from ui_theme import configure_theme, style_text_widget

SOURCE_HEADERS = [
    "渠道属性", "奥莱中类", "产品定义", "货品属性", "线上线下", "办事处", "店仓名称", "店仓",
    "大类", "大小童", "配货季", "货号", "零售数量", "结算金额", "零售吊牌额", "库存数量",
    "库存吊牌额", "销售小票号",
]
REQUIRED_HEADERS = ("渠道属性", "线上线下", "办事处", "店仓名称", "大类", "配货季", "零售数量", "结算金额")
REPORT_HEADERS = [
    "办事处", "店仓名称", "渠道属性", "预算目标", "线上预算目标", "线下预算目标",
    "流水", "线上流水", "线下流水", "预算达成率", "线上预算达成率", "线下预算达成率",
    "预算缺口", "线上缺口", "线下缺口", "旧货目标", "旧货销售额", "旧货达成率",
    "旧货业绩缺口", "旧货占比", "新货销售额", "新货占比", "鞋双数目标", "实际销量",
    "鞋双数达成", "鞋双数缺口", "鞋业绩占比",
]
EXCEL_MAX_ROW = 1_048_576
LogCallback = Callable[[str], None]
ProgressCallback = Callable[[int, int], None]

NAVY, GOLD, AMBER, GREEN, TEAL = "172033", "A4773E", "B36B32", "3D7C59", "287C82"
WARM_BG, CARD, MUTED, BORDER, INPUT = "F4F1EC", "FCFBF9", "74777D", "DDD8CF", "FFF2CC"
WHITE, SOFT_RED, SOFT_GREEN, SOFT_AMBER = "FFFFFF", "FCE8E6", "E4F3E9", "FFF0D6"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _normalise_header(value: object) -> str:
    """移除字段名中的全部空白（含换行、全角空格等）。"""
    return re.sub(r"\s+", "", str(value or "")).strip()


def _has_value(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _emit(callback: LogCallback | None, text: str) -> None:
    if callback is not None:
        callback(text)


def _advance(callback: ProgressCallback | None, value: int, maximum: int) -> None:
    if callback is not None:
        callback(value, maximum)


def _find_header(worksheet) -> tuple[int, dict[str, int]]:
    """在前 100 行中寻找标准字段命中数最多的行；重复字段保留第一列。"""
    wanted = set(SOURCE_HEADERS)
    best_row = 1
    best_mapping: dict[str, int] = {}
    scan_end = min(100, max(1, worksheet.max_row))
    for row in range(1, scan_end + 1):
        mapping: dict[str, int] = {}
        for column in range(1, worksheet.max_column + 1):
            name = _normalise_header(worksheet.cell(row, column).value)
            if name in wanted and name not in mapping:
                mapping[name] = column
        if len(mapping) > len(best_mapping):
            best_row, best_mapping = row, mapping
    return best_row, best_mapping


def _scan_input(workbook, log: LogCallback | None) -> list[tuple[object, int, dict[str, int], int]]:
    scans: list[tuple[object, int, dict[str, int], int]] = []
    failures: list[str] = []
    for worksheet in workbook.worksheets:
        header_row, mapping = _find_header(worksheet)
        missing = [name for name in REQUIRED_HEADERS if name not in mapping]
        _emit(log, f"来源工作表={worksheet.title}；表头行={header_row}；字段映射={mapping}")
        if missing:
            failures.append(f"工作表“{worksheet.title}”缺少必需字段：{'、'.join(missing)}")
            continue
        mapped_columns = tuple(mapping.values())
        last_row = header_row
        for row in range(worksheet.max_row, header_row, -1):
            if any(_has_value(worksheet.cell(row, column).value) for column in mapped_columns):
                last_row = row
                break
        scans.append((worksheet, header_row, mapping, last_row))
    if failures:
        raise ValueError("输入工作簿字段校验失败：\n" + "\n".join(failures))
    return scans


def _style_source(source, copied_rows: int) -> None:
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    source.sheet_view.showGridLines = False
    source.merge_cells("A1:R1")
    source["A1"] = "经营分析数据源"
    source["A1"].fill = _fill(NAVY)
    source["A1"].font = Font(color=WHITE, size=18, bold=True)
    source["A1"].alignment = Alignment(horizontal="left", vertical="center")
    source.row_dimensions[1].height = 34
    source.merge_cells("A2:R2")
    source["A2"] = "汇总输入工作簿全部工作表；经营分析按办事处、店仓名称、渠道属性动态计算，新旧货年度取自配货季前两位。"
    source["A2"].fill = _fill(WARM_BG)
    source["A2"].font = Font(color=MUTED, italic=True)
    source["A2"].alignment = Alignment(vertical="center")
    source.row_dimensions[2].height = 26
    for column, header in enumerate(SOURCE_HEADERS, 1):
        cell = source.cell(4, column, header)
        cell.fill = _fill(GOLD)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in range(5, 5 + copied_rows):
        for column in range(1, len(SOURCE_HEADERS) + 1):
            cell = source.cell(row, column)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row % 2 == 0:
                cell.fill = _fill("F8F5F0")
    source.freeze_panes = "A5"
    source.auto_filter.ref = f"A4:R{max(4, 4 + copied_rows)}"
    widths = [14, 14, 15, 13, 12, 14, 18, 10, 12, 12, 11, 14, 12, 14, 14, 12, 14, 18]
    for column, width in enumerate(widths, 1):
        source.column_dimensions[get_column_letter(column)].width = width


def _merge_card(worksheet, start_column: int, end_column: int, label: str, formula: str, color: str) -> None:
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    worksheet.merge_cells(start_row=4, start_column=start_column, end_row=4, end_column=end_column)
    worksheet.merge_cells(start_row=5, start_column=start_column, end_row=5, end_column=end_column)
    worksheet.cell(4, start_column, label)
    worksheet.cell(5, start_column, formula)
    for row in (4, 5):
        for column in range(start_column, end_column + 1):
            cell = worksheet.cell(row, column)
            cell.fill = _fill(color)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(color=WHITE, bold=True, size=10 if row == 4 else 15)
    worksheet.cell(5, start_column).number_format = "#,##0.00"


def _safe_formula(row: int, expression: str) -> str:
    return f'=IF(COUNTA($A{row}:$C{row})<3,"",{expression})'


def _target_formula(row: int, target_column: str, expression: str) -> str:
    return f'=IF({target_column}{row}="","",IFERROR({expression},0))'


def _build_report(workbook, source_header_map: dict[str, int], formula_end: int,
                  combinations: list[tuple[object, object, object]]) -> tuple[int, int]:
    report = workbook.create_sheet("经营分析")
    report.sheet_view.showGridLines = False
    report.sheet_properties.pageSetUpPr.fitToPage = True
    report.page_setup.fitToWidth = 1
    report.page_setup.fitToHeight = 0
    report.freeze_panes = "D9"
    report.merge_cells("A1:AA1")
    report["A1"] = "经营分析看板"
    report["A1"].fill = _fill(NAVY)
    report["A1"].font = Font(color=WHITE, size=20, bold=True)
    report["A1"].alignment = Alignment(horizontal="left", vertical="center")
    report.row_dimensions[1].height = 40
    report.merge_cells("A2:AA2")
    report["A2"] = "自动公式模板  ·  当前年度由 TODAY() 动态判断  ·  黄色单元格为手工目标输入区"
    report["A2"].fill = _fill(WARM_BG)
    report["A2"].font = Font(color=MUTED, italic=True, size=10)
    report["A2"].alignment = Alignment(horizontal="left", vertical="center")
    report.row_dimensions[2].height = 24

    detail_start = 9
    template_rows = max(40, len(combinations) + 30)
    detail_end = detail_start + template_rows - 1
    cards = [
        (1, 4, "总流水", f"=SUM(G{detail_start}:G{detail_end})", NAVY),
        (5, 8, "线上流水", f"=SUM(H{detail_start}:H{detail_end})", GOLD),
        (9, 12, "线下流水", f"=SUM(I{detail_start}:I{detail_end})", GOLD),
        (13, 16, "旧货销售额", f"=SUM(Q{detail_start}:Q{detail_end})", AMBER),
        (17, 20, "新货销售额", f"=SUM(U{detail_start}:U{detail_end})", GREEN),
        (21, 24, "鞋类实际销量", f"=SUM(X{detail_start}:X{detail_end})", TEAL),
        (25, 27, "有效组合", f'=COUNTIFS(A{detail_start}:A{detail_end},"<>",B{detail_start}:B{detail_end},"<>",C{detail_start}:C{detail_end},"<>")', NAVY),
    ]
    for card in cards:
        _merge_card(report, *card)
    report["U5"].number_format = "#,##0"
    report["Y5"].number_format = "#,##0"
    report.row_dimensions[4].height = 23
    report.row_dimensions[5].height = 32

    groups = [
        (1, 3, "门店信息", NAVY),
        (4, 15, "预算与达成", GOLD),
        (16, 20, "旧货表现 · 配货季早于当前年份", AMBER),
        (21, 22, "新货表现 · 当前年份", GREEN),
        (23, 27, "鞋类表现 · 仅统计线下鞋类", TEAL),
    ]
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_colors: dict[int, str] = {}
    for start_column, end_column, title, color in groups:
        report.merge_cells(start_row=7, start_column=start_column, end_row=7, end_column=end_column)
        heading = report.cell(7, start_column, title)
        heading.font = Font(color=WHITE, bold=True, size=11)
        heading.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(start_column, end_column + 1):
            section_colors[column] = color
            report.cell(7, column).fill = _fill(color)
            report.cell(7, column).border = border
    report.row_dimensions[7].height = 26
    for column, header in enumerate(REPORT_HEADERS, 1):
        cell = report.cell(8, column, header)
        cell.fill = _fill(section_colors[column])
        cell.font = Font(color=WHITE, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    report.row_dimensions[8].height = 44

    def source_range(field: str) -> str:
        letter = get_column_letter(source_header_map[field])
        return f"'数据源'!${letter}$5:${letter}${formula_end}"

    office = source_range("办事处")
    store = source_range("店仓名称")
    channel = source_range("渠道属性")
    online = source_range("线上线下")
    category = source_range("大类")
    season = source_range("配货季")
    quantity = source_range("零售数量")
    amount = source_range("结算金额")

    for index in range(template_rows):
        row = detail_start + index
        if index < len(combinations):
            for column, value in enumerate(combinations[index], 1):
                report.cell(row, column, value)
        criteria = f"{office},$A{row},{store},$B{row},{channel},$C{row}"
        report.cell(row, 7, _safe_formula(row, f"SUMIFS({amount},{criteria})"))
        report.cell(row, 8, _safe_formula(row, f'SUMIFS({amount},{criteria},{online},"线上")'))
        report.cell(row, 9, _safe_formula(row, f'SUMIFS({amount},{criteria},{online},"线下")'))
        report.cell(row, 10, _target_formula(row, "D", f"G{row}/D{row}"))
        report.cell(row, 11, _target_formula(row, "E", f"H{row}/E{row}"))
        report.cell(row, 12, _target_formula(row, "F", f"I{row}/F{row}"))
        report.cell(row, 13, f'=IF(D{row}="","",G{row}-D{row})')
        report.cell(row, 14, f'=IF(E{row}="","",H{row}-E{row})')
        report.cell(row, 15, f'=IF(F{row}="","",I{row}-F{row})')
        old_sales = (
            f"SUMPRODUCT(({office}=$A{row})*({store}=$B{row})*({channel}=$C{row})*"
            f'({online}="线下")*(IFERROR(--LEFT({season},2),999)<MOD(YEAR(TODAY()),100))*'
            f"IFERROR({amount},0))"
        )
        report.cell(row, 17, _safe_formula(row, old_sales))
        report.cell(row, 18, _target_formula(row, "P", f"Q{row}/P{row}"))
        report.cell(row, 19, f'=IF(P{row}="","",Q{row}-P{row})')
        report.cell(row, 20, f'=IF(OR(I{row}="",I{row}=0),"",Q{row}/I{row})')
        new_sales = (
            f"SUMPRODUCT(({office}=$A{row})*({store}=$B{row})*({channel}=$C{row})*"
            f'({online}="线下")*(IFERROR(--LEFT({season},2),999)=MOD(YEAR(TODAY()),100))*'
            f"IFERROR({amount},0))"
        )
        report.cell(row, 21, _safe_formula(row, new_sales))
        report.cell(row, 22, f'=IF(OR(I{row}="",I{row}=0),"",U{row}/I{row})')
        shoe_quantity = f'SUMIFS({quantity},{criteria},{online},"线下",{category},"鞋类")'
        report.cell(row, 24, _safe_formula(row, shoe_quantity))
        report.cell(row, 25, _target_formula(row, "W", f"X{row}/W{row}"))
        report.cell(row, 26, f'=IF(W{row}="","",X{row}-W{row})')
        shoe_amount = f'SUMIFS({amount},{criteria},{online},"线下",{category},"鞋类")'
        report.cell(row, 27, f'=IF(OR(I{row}="",I{row}=0),"",{shoe_amount}/I{row})')

    input_columns = {4, 5, 6, 16, 23}
    rate_columns = {10, 11, 12, 18, 20, 22, 25, 27}
    amount_columns = {4, 5, 6, 7, 8, 9, 13, 14, 15, 16, 17, 19, 21}
    quantity_columns = {23, 24, 26}
    for row in range(detail_start, detail_end + 1):
        report.row_dimensions[row].height = 24
        for column in range(1, 28):
            cell = report.cell(row, column)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if column <= 3 else "right", vertical="center")
            cell.fill = _fill("F8F5F0" if row % 2 == 0 else CARD)
            if column in input_columns:
                cell.fill = _fill(INPUT)
                cell.protection = Protection(locked=False)
            if column in rate_columns:
                cell.number_format = "0.00%"
            elif column in amount_columns:
                cell.number_format = "#,##0.00;[Red]-#,##0.00"
            elif column in quantity_columns:
                cell.number_format = "#,##0;[Red]-#,##0"
    report.auto_filter.ref = f"A8:AA{detail_end}"
    report.column_dimensions["A"].width = 14
    report.column_dimensions["B"].width = 18
    report.column_dimensions["C"].width = 14
    for column in range(4, 28):
        report.column_dimensions[get_column_letter(column)].width = 13 if column in rate_columns else 14

    for range_string in (f"J{detail_start}:L{detail_end}", f"R{detail_start}:R{detail_end}", f"Y{detail_start}:Y{detail_end}"):
        first = range_string.split(":", 1)[0]
        report.conditional_formatting.add(range_string, FormulaRule(
            formula=[f'AND({first}<>"",{first}>=1)'], fill=_fill(SOFT_GREEN), font=Font(color="246B45", bold=True)))
        report.conditional_formatting.add(range_string, FormulaRule(
            formula=[f'AND({first}<>"",{first}>=0.8,{first}<1)'], fill=_fill(SOFT_AMBER), font=Font(color="8A5A16", bold=True)))
        report.conditional_formatting.add(range_string, FormulaRule(
            formula=[f'AND({first}<>"",{first}<0.8)'], fill=_fill(SOFT_RED), font=Font(color="9C2F2F", bold=True)))
    for range_string in (f"M{detail_start}:O{detail_end}", f"S{detail_start}:S{detail_end}", f"Z{detail_start}:Z{detail_end}"):
        first = range_string.split(":", 1)[0]
        report.conditional_formatting.add(range_string, FormulaRule(
            formula=[f'AND({first}<>"",{first}>=0)'], fill=_fill(SOFT_GREEN), font=Font(color="246B45")))
        report.conditional_formatting.add(range_string, FormulaRule(
            formula=[f'AND({first}<>"",{first}<0)'], fill=_fill(SOFT_RED), font=Font(color="9C2F2F")))
    for range_string, color in (
        (f"T{detail_start}:T{detail_end}", AMBER),
        (f"V{detail_start}:V{detail_end}", GREEN),
        (f"AA{detail_start}:AA{detail_end}", TEAL),
    ):
        report.conditional_formatting.add(range_string, DataBarRule(
            start_type="num", start_value=0, end_type="num", end_value=1, color=color, showValue=True))
    report.print_title_rows = "1:8"
    report.sheet_view.zoomScale = 80
    return template_rows, detail_end


def _self_check(output_path: Path, formula_end: int, log: LogCallback | None) -> int:
    check = load_workbook(output_path, data_only=False)
    try:
        required_sheets = {"数据源", "经营分析"}
        if not required_sheets.issubset(check.sheetnames):
            raise AssertionError("输出工作簿缺少“数据源”或“经营分析”工作表")
        source = check["数据源"]
        report = check["经营分析"]
        scanned_headers = {
            _normalise_header(source.cell(4, column).value): column
            for column in range(1, source.max_column + 1)
            if _normalise_header(source.cell(4, column).value)
        }
        if any(name not in scanned_headers for name in SOURCE_HEADERS):
            raise AssertionError("数据源第4行标准表头不完整")
        for address in ("G9", "Q9", "X9"):
            if report[address].data_type != "f":
                raise AssertionError(f"关键单元格 {address} 不是公式")
        formulas: list[str] = []
        for worksheet in check.worksheets:
            formulas.extend(
                str(cell.value)
                for row in worksheet.iter_rows()
                for cell in row
                if cell.data_type == "f"
            )
        if not formulas:
            raise AssertionError("输出工作簿未生成公式")
        if any("#REF!" in formula.upper() for formula in formulas):
            raise AssertionError("公式中存在 #REF! 引用错误")
        if any("SOURCEDATA[" in formula.upper() for formula in formulas):
            raise AssertionError("公式中存在禁止的 SourceData[ 结构化引用")

        allowed_letters = {get_column_letter(column) for column in scanned_headers.values()}
        pattern = re.compile(r"'数据源'!\$([A-Z]{1,3})\$(\d+):\$([A-Z]{1,3})\$(\d+)")
        source_references = [match.groups() for formula in formulas for match in pattern.finditer(formula)]
        if not source_references:
            raise AssertionError("经营分析公式未引用数据源普通区域")
        for start_letter, start_row, end_letter, end_row in source_references:
            if (start_letter != end_letter or start_letter not in allowed_letters
                    or int(start_row) != 5 or int(end_row) != formula_end):
                raise AssertionError(f"公式数据源引用与第4行表头映射不一致：{start_letter}{start_row}:{end_letter}{end_row}")
        expected = {
            "G9": "结算金额",
            "H9": "线上线下",
            "Q9": "配货季",
            "X9": "零售数量",
            "AA9": "大类",
        }
        for address, field in expected.items():
            letter = get_column_letter(scanned_headers[field])
            token = f"'数据源'!${letter}$5:${letter}${formula_end}"
            if token not in str(report[address].value):
                raise AssertionError(f"公式 {address} 未按第4行表头定位字段“{field}”")
        if report.freeze_panes != "D9":
            raise AssertionError("经营分析冻结窗格不是 D9")
        _emit(log, f"自检结果=通过；公式数={len(formulas)}；数据源普通区域引用数={len(source_references)}")
        return len(formulas)
    finally:
        check.close()


def build_analysis_template(input_path, output_path, log=None, progress=None) -> Path:
    """读取原始 Excel，生成含标准数据源与经营分析看板的新 XLSX。"""
    from weekly_report import read_table_content

    source_path = Path(input_path).expanduser()
    target_path = Path(output_path).expanduser()
    if source_path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError("输入文件必须是 .xlsx 或 .xls 格式")
    if target_path.suffix.lower() != ".xlsx":
        raise ValueError("输出文件必须是 .xlsx 格式")
    if not source_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{source_path}")
    if not target_path.parent.is_dir():
        raise FileNotFoundError(f"输出目录不存在：{target_path.parent}")
    if source_path.resolve() == target_path.resolve():
        raise ValueError("输出文件不能覆盖输入文件")

    _emit(log, f"开始读取：{source_path}")
    rows, number_formats = read_table_content(source_path)
    best_header_index = -1
    best_mapping: dict[str, int] = {}
    for row_index, row in enumerate(rows[:100]):
        mapping: dict[str, int] = {}
        for column_index, value in enumerate(row):
            name = _normalise_header(value)
            if name in SOURCE_HEADERS and name not in mapping:
                mapping[name] = column_index
        if len(mapping) > len(best_mapping):
            best_header_index, best_mapping = row_index, mapping
        if all(field in mapping for field in REQUIRED_HEADERS):
            best_header_index, best_mapping = row_index, mapping
            break
    missing = [field for field in REQUIRED_HEADERS if field not in best_mapping]
    if missing:
        raise ValueError(
            f"{source_path.name} 前100行无法识别完整表头（已匹配 "
            f"{len(REQUIRED_HEADERS) - len(missing)}/{len(REQUIRED_HEADERS)} 个必需字段），"
            f"缺少：{'、'.join(missing)}"
        )
    _emit(
        log,
        f"表头行={best_header_index + 1}；字段映射="
        f"{ {field: column + 1 for field, column in best_mapping.items()} }",
    )
    _advance(progress, 1, 4)

    standardized_rows: list[tuple[list[object], list[str]]] = []
    combinations: list[tuple[object, object, object]] = []
    seen: set[tuple[object, object, object]] = set()
    for row_index in range(best_header_index + 1, len(rows)):
        input_row = rows[row_index]
        values = [
            input_row[best_mapping[field]]
            if field in best_mapping and best_mapping[field] < len(input_row)
            else None
            for field in SOURCE_HEADERS
        ]
        if not any(_has_value(value) for value in values):
            continue
        input_formats = number_formats[row_index] if number_formats and row_index < len(number_formats) else []
        formats = [
            input_formats[best_mapping[field]]
            if field in best_mapping and best_mapping[field] < len(input_formats)
            else "General"
            for field in SOURCE_HEADERS
        ]
        standardized_rows.append((values, formats))
        combo = (values[5], values[6], values[0])
        if all(_has_value(value) for value in combo) and combo not in seen:
            seen.add(combo)
            combinations.append(combo)
    if len(standardized_rows) > EXCEL_MAX_ROW - 4:
        raise ValueError(f"输出数据行数 {len(standardized_rows)} 超过 Excel 上限 {EXCEL_MAX_ROW - 4}")
    _emit(
        log,
        f"原始数据行数={max(0, len(rows) - best_header_index - 1)}；"
        f"有效输出行数={len(standardized_rows)}；唯一组合数={len(combinations)}",
    )
    _advance(progress, 2, 4)

    output_workbook = Workbook()
    try:
        source = output_workbook.active
        source.title = "数据源"
        for row_index, (values, formats) in enumerate(standardized_rows, 5):
            for column_index, value in enumerate(values, 1):
                cell = source.cell(row_index, column_index, value)
                cell.number_format = formats[column_index - 1] or "General"
        _style_source(source, len(standardized_rows))
        source_header_map = {
            _normalise_header(source.cell(4, column).value): column
            for column in range(1, source.max_column + 1)
            if _normalise_header(source.cell(4, column).value)
        }
        formula_end = max(5, 4 + len(standardized_rows))
        template_rows, detail_end = _build_report(
            output_workbook, source_header_map, formula_end, combinations
        )
        output_workbook.calculation.fullCalcOnLoad = True
        output_workbook.calculation.forceFullCalc = True
        output_workbook.calculation.calcMode = "auto"
        _emit(
            log,
            f"经营分析明细行数={template_rows}（结束行={detail_end}）；"
            f"公式数据区域=数据源第5至{formula_end}行",
        )
        _advance(progress, 3, 4)
        output_workbook.save(target_path)
    finally:
        output_workbook.close()

    try:
        formula_count = _self_check(target_path, formula_end, log)
    except Exception:
        try:
            target_path.unlink()
        except OSError:
            pass
        raise
    _advance(progress, 4, 4)
    result = target_path.resolve()
    _emit(log, f"生成完成：{result}；公式数={formula_count}")
    return result


class AnalysisTemplateWindow:
    """独立经营分析模板窗口；只创建单个 Toplevel。"""

    def __init__(self, parent: tk.Misc):
        self.window = tk.Toplevel(parent)
        configure_theme(self.window)
        self.window.title("经营分析模板")
        self.window.geometry("920x700")
        self.window.minsize(820, 620)
        self.input_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.messages: queue.Queue[dict[str, object]] = queue.Queue()
        self._build_ui()
        self._poll_messages()

    def _build_ui(self) -> None:
        container = ttk.Frame(self.window, padding=(20, 16), style="App.TFrame")
        container.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(container, padding=(24, 17), style="Header.TFrame")
        header.pack(fill=tk.X, pady=(0, 12))
        copy = ttk.Frame(header, style="Header.TFrame")
        copy.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(copy, text="经营分析模板", style="HeaderTitle.TLabel").pack(anchor=tk.W)
        ttk.Label(
            copy,
            text="识别多工作表数据 · 自动生成动态经营看板",
            style="HeaderSubtitle.TLabel",
        ).pack(anchor=tk.W, pady=(4, 0))
        self.generate_button = ttk.Button(
            header,
            text="生成模板",
            command=self._start_generation,
            style="Header.Primary.TButton",
        )
        self.generate_button.pack(side=tk.RIGHT, padx=(18, 0))

        file_card = ttk.LabelFrame(
            container,
            text="01  选择输入文件",
            padding=(16, 14),
            style="Card.TLabelframe",
        )
        file_card.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(
            file_card,
            text="支持 .xlsx / .xls；扫描前 100 行按字段名识别表头，输出始终另存为新的 .xlsx。",
            style="CardText.TLabel",
        ).grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 10))
        ttk.Label(file_card, text="原始文件", style="CardText.TLabel").grid(
            row=1, column=0, sticky=tk.W, padx=(0, 10)
        )
        ttk.Entry(
            file_card,
            textvariable=self.input_var,
            state="readonly",
            style="Modern.TEntry",
        ).grid(row=1, column=1, sticky=tk.EW, padx=(0, 10))
        ttk.Button(
            file_card,
            text="选择文件",
            command=self._choose_input,
            style="Secondary.TButton",
        ).grid(row=1, column=2)
        ttk.Label(file_card, text="保存路径", style="CardText.TLabel").grid(
            row=2, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0)
        )
        ttk.Entry(
            file_card,
            textvariable=self.output_var,
            state="readonly",
            style="Modern.TEntry",
        ).grid(row=2, column=1, sticky=tk.EW, padx=(0, 10), pady=(10, 0))
        ttk.Button(
            file_card,
            text="选择路径",
            command=self._choose_output,
            style="Secondary.TButton",
        ).grid(row=2, column=2, pady=(10, 0))
        file_card.columnconfigure(1, weight=1)

        status_card = ttk.Frame(container, padding=(16, 13), style="Card.TFrame")
        status_card.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(status_card, text="生成进度", style="CardTitle.TLabel").pack(side=tk.LEFT, padx=(0, 14))
        self.progress_bar = ttk.Progressbar(
            status_card,
            mode="determinate",
            style="Business.Horizontal.TProgressbar",
        )
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 12))
        self.progress_label = ttk.Label(status_card, text="就绪", style="Status.TLabel")
        self.progress_label.pack(side=tk.RIGHT)

        log_card = ttk.LabelFrame(
            container,
            text="02  处理日志",
            padding=12,
            style="Card.TLabelframe",
        )
        log_card.pack(fill=tk.BOTH, expand=True)
        toolbar = ttk.Frame(log_card, style="Card.TFrame")
        toolbar.pack(fill=tk.X, pady=(0, 7))
        ttk.Label(
            toolbar,
            text="记录表头、字段映射、数据行、组合、公式及自检结果",
            style="CardText.TLabel",
        ).pack(side=tk.LEFT)
        ttk.Button(
            toolbar,
            text="清空日志",
            command=self._clear_log,
            style="Ghost.TButton",
            padding=(8, 4),
        ).pack(side=tk.RIGHT)
        log_body = ttk.Frame(log_card, style="Card.TFrame")
        log_body.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(log_body, height=18, wrap=tk.WORD, state=tk.DISABLED)
        style_text_widget(self.log_text)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(
            log_body,
            orient=tk.VERTICAL,
            command=self.log_text.yview,
            style="Modern.Vertical.TScrollbar",
        )
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        footer = ttk.Frame(container, style="App.TFrame")
        footer.pack(fill=tk.X, pady=(12, 0))
        ttk.Label(
            footer,
            text="黄色列为目标录入区；模板打开后将自动重算全部公式。",
            style="Muted.TLabel",
        ).pack(side=tk.LEFT)
        ttk.Button(
            footer,
            text="关闭窗口",
            command=self.window.destroy,
            style="Ghost.TButton",
        ).pack(side=tk.RIGHT)

    def _choose_input(self) -> None:
        selected = filedialog.askopenfilename(
            parent=self.window,
            title="选择经营数据文件",
            filetypes=[("Excel 工作簿", "*.xlsx *.xls")],
        )
        if selected:
            self.input_var.set(selected)
            input_path = Path(selected)
            self.output_var.set(str(input_path.with_name(f"{input_path.stem}_经营分析模板.xlsx")))
            self._append_log(f"已选择输入文件：{selected}")

    def _choose_output(self) -> None:
        input_name = self.input_var.get().strip()
        input_path = Path(input_name) if input_name else Path.cwd() / "经营分析模板.xlsx"
        selected = filedialog.asksaveasfilename(
            parent=self.window,
            title="保存经营分析模板",
            initialdir=str(input_path.parent),
            initialfile=f"{input_path.stem}_经营分析模板.xlsx" if input_name else input_path.name,
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if selected:
            self.output_var.set(selected)
            self._append_log(f"已选择保存路径：{selected}")

    def _start_generation(self) -> None:
        input_name = self.input_var.get().strip()
        output_name = self.output_var.get().strip()
        if not input_name:
            messagebox.showwarning("缺少输入文件", "请先选择一个 .xlsx 或 .xls 文件。", parent=self.window)
            return
        if not output_name:
            messagebox.showwarning("缺少保存路径", "请先选择输出 .xlsx 的保存路径。", parent=self.window)
            return
        if Path(input_name).resolve() == Path(output_name).resolve():
            messagebox.showwarning("保存路径无效", "输出文件不能覆盖输入文件。", parent=self.window)
            return
        self.generate_button.configure(state=tk.DISABLED)
        self.progress_bar.configure(value=0, maximum=1)
        self.progress_label.configure(text="准备中")
        self._append_log("开始生成经营分析模板……")
        threading.Thread(
            target=self._generation_worker,
            args=(input_name, output_name),
            daemon=True,
        ).start()

    def _generation_worker(self, input_name: str, output_name: str) -> None:
        try:
            result = build_analysis_template(
                input_name,
                output_name,
                log=lambda text: self.messages.put({"type": "log", "text": text}),
                progress=lambda value, maximum: self.messages.put(
                    {"type": "progress", "value": value, "maximum": maximum}
                ),
            )
            self.messages.put({"type": "done", "result": result})
        except Exception as error:
            self.messages.put({"type": "error", "text": str(error)})


    def _poll_messages(self) -> None:
        try:
            if not self.window.winfo_exists():
                return
        except tk.TclError:
            return
        try:
            while True:
                message = self.messages.get_nowait()
                message_type = message.get("type")
                if message_type == "log":
                    self._append_log(str(message.get("text", "")))
                elif message_type == "progress":
                    value = int(message.get("value", 0))
                    maximum = max(1, int(message.get("maximum", 1)))
                    self.progress_bar.configure(value=value, maximum=maximum)
                    self.progress_label.configure(text=f"{value}/{maximum}")
                elif message_type == "done":
                    self._handle_done(Path(str(message["result"])))
                elif message_type == "error":
                    self._handle_error(str(message.get("text", "未知错误")))
        except queue.Empty:
            pass
        self.window.after(100, self._poll_messages)

    def _append_log(self, text: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state=tk.NORMAL)
        for line in text.splitlines() or [""]:
            self.log_text.insert(tk.END, f"[{timestamp}] {line}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _clear_log(self) -> None:
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _handle_done(self, result: Path) -> None:
        self.generate_button.configure(state=tk.NORMAL)
        self.progress_label.configure(text="已完成")
        self._append_log(f"模板已保存：{result}")
        messagebox.showinfo("生成完成", f"经营分析模板已生成：\n{result}", parent=self.window)

    def _handle_error(self, text: str) -> None:
        self.generate_button.configure(state=tk.NORMAL)
        self.progress_label.configure(text="生成失败")
        self._append_log(f"生成失败：{text}")
        messagebox.showerror("生成失败", text, parent=self.window)


__all__ = ["build_analysis_template", "AnalysisTemplateWindow"]
