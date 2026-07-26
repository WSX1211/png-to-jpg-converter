#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""周报合并：读取 Excel/CSV，并将各时间来源纵向追加到主表。"""

from __future__ import annotations

import csv
import queue
import re
import threading
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from ui_theme import configure_theme, style_text_widget


DEFAULT_BASE_FIELDS = (
    "仓店类型", "办事处", "店铺类别", "仓店名称", "仓店代码", "销售截止日期", "货号",
    "商品名称", "牌价", "大类", "系列", "大小童", "性别", "配货季",
)
DEFAULT_METRICS = ("零售数量", "结算金额", "零售吊牌额", "财务存", "财务存牌价")
SOURCE_ROLES = ("同期月", "本周", "上周", "同期周")
SUPPORTED_SUFFIXES = {".xlsx", ".xls", ".csv"}


@dataclass
class TableData:
    path: Path
    rows: list[list[object]]
    header_index: int
    header_map: dict[str, list[int]]
    number_formats: list[list[str]] | None = None
    summary_row_index: int | None = None


@dataclass
class MergeResult:
    output_path: Path
    source_counts: dict[str, int]


def normalize_header(value: object) -> str:
    """统一表头比较形式，忽略空白和换行。"""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def parse_field_text(text: str) -> list[str]:
    """解析按行、逗号或制表符输入的动态字段。"""
    parts = re.split(r"[\n,，\t]+", text)
    fields: list[str] = []
    seen: set[str] = set()
    for part in parts:
        field = part.strip()
        normalized = normalize_header(field)
        if normalized and normalized not in seen:
            fields.append(field)
            seen.add(normalized)
    return fields


def _read_xlsx(path: Path) -> tuple[list[list[object]], list[list[str]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows: list[list[object]] = []
        formats: list[list[str]] = []
        for cells in sheet.iter_rows():
            rows.append([cell.value for cell in cells])
            formats.append([cell.number_format or "General" for cell in cells])
        return rows, formats
    finally:
        workbook.close()


def _read_xls(path: Path) -> tuple[list[list[object]], list[list[str]]]:
    import xlrd

    workbook = xlrd.open_workbook(path, formatting_info=True)
    sheet = workbook.sheet_by_index(0)
    rows: list[list[object]] = []
    formats: list[list[str]] = []
    for row_index in range(sheet.nrows):
        row: list[object] = []
        row_formats: list[str] = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode))
            else:
                row.append(cell.value)
            try:
                format_key = workbook.xf_list[cell.xf_index].format_key
                row_formats.append(workbook.format_map[format_key].format_str or "General")
            except (AttributeError, IndexError, KeyError):
                row_formats.append("General")
        rows.append(row)
        formats.append(row_formats)
    return rows, formats


def _coerce_csv_cell(value: str) -> object:
    """保守恢复 CSV 数值；前导零代码和超长整数保持文本。"""
    text = value.strip()
    if not text:
        return None
    compact = text.replace(",", "") if re.fullmatch(r"[+-]?\d{1,3}(?:,\d{3})+(?:\.\d+)?", text) else text
    if re.fullmatch(r"[+-]?0\d+", compact):
        return text
    if re.fullmatch(r"[+-]?\d+", compact):
        digits = compact.lstrip("+-")
        return int(compact) if len(digits) <= 15 else text
    if re.fullmatch(r"[+-]?(?:\d+\.\d*|\.\d+)(?:[eE][+-]?\d+)?", compact):
        return float(compact)
    if re.fullmatch(r"[+-]?\d+[eE][+-]?\d+", compact):
        return float(compact)
    return value


def _read_csv(path: Path) -> tuple[list[list[object]], None]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                sample = file.read(8192)
                file.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                except csv.Error:
                    dialect = csv.excel
                rows = [list(row) for row in csv.reader(file, dialect)]
                return rows, None
        except UnicodeDecodeError as error:
            last_error = error
    raise ValueError(f"无法识别 CSV 编码：{path.name}") from last_error


def read_table_content(path: str | Path) -> tuple[list[list[object]], list[list[str]] | None]:
    file_path = Path(path)
    if not file_path.exists():
        raise ValueError(f"文件不存在：{file_path}")
    suffix = file_path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"不支持的文件格式：{file_path.name}")
    if suffix == ".xlsx":
        rows, formats = _read_xlsx(file_path)
    elif suffix == ".xls":
        rows, formats = _read_xls(file_path)
    else:
        rows, formats = _read_csv(file_path)
    if not rows:
        raise ValueError(f"文件没有数据：{file_path.name}")
    return rows, formats


def read_rows(path: str | Path) -> list[list[object]]:
    """兼容原调用，仅返回单元格值。"""
    return read_table_content(path)[0]


def build_header_map(row: Iterable[object]) -> dict[str, list[int]]:
    header_map: dict[str, list[int]] = defaultdict(list)
    for index, value in enumerate(row):
        normalized = normalize_header(value)
        if normalized:
            header_map[normalized].append(index)
    return dict(header_map)


def detect_header(rows: list[list[object]], fields: list[str], file_name: str) -> tuple[int, dict[str, list[int]]]:
    required = {normalize_header(field) for field in fields}
    best_index = -1
    best_map: dict[str, list[int]] = {}
    best_score = -1
    for index, row in enumerate(rows[:100]):
        header_map = build_header_map(row)
        score = len(required.intersection(header_map))
        if score > best_score:
            best_index, best_map, best_score = index, header_map, score
        if required.issubset(header_map):
            return index, header_map
    missing = sorted(required.difference(best_map))
    matched = f"已匹配 {best_score}/{len(required)} 个字段"
    raise ValueError(f"{file_name} 无法识别完整表头（{matched}），缺少：{'、'.join(missing)}")


def _coerce_csv_numeric_columns(
    rows: list[list[object]], header_index: int, header_map: dict[str, list[int]], metrics: list[str]
) -> None:
    numeric_fields = set(DEFAULT_METRICS) | {"牌价", "行位"} | set(metrics)
    numeric_indexes = {
        index
        for field in numeric_fields
        for index in header_map.get(normalize_header(field), [])
    }
    for row in rows[header_index + 1:]:
        for index in numeric_indexes:
            if index < len(row) and isinstance(row[index], str):
                row[index] = _coerce_csv_cell(row[index])


def load_table(path: str | Path, base_fields: list[str], metrics: list[str]) -> TableData:
    file_path = Path(path)
    rows, number_formats = read_table_content(file_path)
    header_index, header_map = detect_header(rows, base_fields + metrics, file_path.name)
    if file_path.suffix.lower() == ".csv":
        _coerce_csv_numeric_columns(rows, header_index, header_map, metrics)
    table = TableData(file_path, rows, header_index, header_map, number_formats)
    table.summary_row_index = _find_final_summary_row(table, base_fields, metrics)
    return table


def _column_index(table: TableData, field: str, use_last: bool) -> int:
    indexes = table.header_map[normalize_header(field)]
    return indexes[-1] if use_last else indexes[0]


def _is_nonempty(value: object) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _has_values(row: list[object], indexes: Iterable[int]) -> bool:
    return any(index < len(row) and _is_nonempty(row[index]) for index in indexes)


def _cell_value(row: list[object], index: int) -> object:
    return row[index] if index < len(row) else None


def _cell_number_format(table: TableData, row_index: int, column_index: int) -> str | None:
    if table.number_formats is None or row_index >= len(table.number_formats):
        return None
    row_formats = table.number_formats[row_index]
    return row_formats[column_index] if column_index < len(row_formats) else None


def _write_value(sheet, row: int, column: int, value: object, number_format: str | None) -> None:
    cell = sheet.cell(row, column, value)
    if number_format and number_format.lower() != "general":
        cell.number_format = number_format


def _last_nonempty_column(row: list[object]) -> int:
    for index in range(len(row) - 1, -1, -1):
        if _is_nonempty(row[index]):
            return index + 1
    return 0


def _last_content_row(rows: list[list[object]], excluded_index: int | None = None) -> int:
    for index in range(len(rows) - 1, -1, -1):
        if index != excluded_index and any(_is_nonempty(value) for value in rows[index]):
            return index + 1
    return 1


def _find_final_summary_row(table: TableData, base_fields: list[str], metrics: list[str]) -> int | None:
    base_indexes = [_column_index(table, field, False) for field in base_fields]
    metric_indexes = [_column_index(table, metric, True) for metric in metrics]
    relevant_indexes = base_indexes + metric_indexes
    final_index: int | None = None
    for index in range(len(table.rows) - 1, table.header_index, -1):
        if _has_values(table.rows[index], relevant_indexes):
            final_index = index
            break
    if final_index is None:
        return None

    row = table.rows[final_index]
    labels = []
    for column_index in base_indexes:
        value = _cell_value(row, column_index)
        if isinstance(value, str):
            labels.append(re.sub(r"[\s:：]+", "", value).lower())
    summary_words = ("合计", "总计", "汇总", "小计")
    has_summary_label = any(
        label in {"total", "grandtotal"}
        or (len(label) <= 12 and any(label.endswith(word) for word in summary_words))
        for label in labels
    )
    has_metric_value = _has_values(row, metric_indexes)
    return final_index if has_summary_label and has_metric_value else None


def _effective_data_rows(
    table: TableData, base_fields: list[str], metrics: list[str]
) -> list[tuple[int, list[object]]]:
    indexes = (
        [_column_index(table, field, False) for field in base_fields]
        + [_column_index(table, metric, True) for metric in metrics]
    )
    return [
        (index, row)
        for index, row in enumerate(table.rows[table.header_index + 1:], start=table.header_index + 1)
        if index != table.summary_row_index and _has_values(row, indexes)
    ]


def _copy_header_style(sheet, source_column: int, target_column: int, header_row: int) -> None:
    source = sheet.cell(header_row, source_column)
    target = sheet.cell(header_row, target_column)
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.fill:
        target.fill = copy(source.fill)
    if source.font:
        target.font = copy(source.font)
    if source.border:
        target.border = copy(source.border)


def _create_output_workbook(main: TableData):
    from openpyxl import Workbook, load_workbook

    if main.path.suffix.lower() == ".xlsx":
        workbook = load_workbook(main.path, data_only=False)
        return workbook, workbook.active
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "周报合并"
    for row_index, row in enumerate(main.rows, start=1):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row_index, column_index, value)
    return workbook, sheet


def merge_reports(
    main_path: str | Path,
    source_paths: dict[str, str | Path],
    base_fields: list[str],
    metrics: list[str],
    output_path: str | Path,
    log: Callable[[str], None] | None = None,
    progress: Callable[[int, int], None] | None = None,
) -> MergeResult:
    """保留主表主体，排除末尾汇总，并按角色纵向追加来源数据。"""
    logger = log or (lambda _message: None)
    if not base_fields:
        raise ValueError("基础字段不能为空")
    if not metrics:
        raise ValueError("追加指标不能为空")
    selected_sources = [(role, Path(source_paths[role])) for role in SOURCE_ROLES if source_paths.get(role)]
    if not selected_sources:
        raise ValueError("请至少选择一个追加文件")

    output = Path(output_path)
    input_paths = [Path(main_path), *(path for _, path in selected_sources)]
    if any(output.resolve() == path.resolve() for path in input_paths):
        raise ValueError("输出文件不能覆盖任何输入文件")

    logger("=" * 72)
    logger("周报合并任务配置")
    logger(f"主文件路径：{Path(main_path).resolve()}")
    for role, path in selected_sources:
        logger(f"{role}路径：{path.resolve()}")
    logger(f"输出路径：{output.resolve()}")
    logger(f"基础字段（{len(base_fields)} 个）：{'、'.join(base_fields)}")
    logger(f"追加指标（{len(metrics)} 个）：{'、'.join(metrics)}")
    logger(f"已选追加角色（{len(selected_sources)} 个）：{'、'.join(role for role, _ in selected_sources)}")

    def log_table_profile(
        label: str,
        table: TableData,
        data_rows: list[tuple[int, list[object]]],
    ) -> None:
        header_columns = _last_nonempty_column(table.rows[table.header_index])
        content_columns = max((_last_nonempty_column(row) for row in table.rows), default=0)
        summary = (
            f"已识别并去掉第 {table.summary_row_index + 1} 行汇总"
            if table.summary_row_index is not None else "末行不是明确汇总，已保留"
        )
        logger(
            f"{label} {table.path.name}：原始读取 {len(table.rows)} 行 × {content_columns} 个含值列；"
            f"有效数据 {len(data_rows)} 行 × {header_columns} 列；表头第 {table.header_index + 1} 行；{summary}"
        )
        base_matches = []
        for field in base_fields:
            positions = table.header_map[normalize_header(field)]
            base_matches.append(f"{field}=第{positions[0] + 1}列")
        logger(f"{label}基础字段匹配：{'；'.join(base_matches)}")

        metric_matches = []
        for metric in metrics:
            positions = table.header_map[normalize_header(metric)]
            detail = f"{metric}=第{positions[-1] + 1}列"
            if len(positions) > 1:
                detail += f"（发现同名 {len(positions)} 列，取最后一列）"
            metric_matches.append(detail)
        logger(f"{label}指标匹配：{'；'.join(metric_matches)}")

    logger("[读取] 正在读取主文件……")
    main = load_table(main_path, base_fields, metrics)
    main_data_rows = _effective_data_rows(main, base_fields, metrics)
    main_columns = _last_nonempty_column(main.rows[main.header_index])
    log_table_profile("主文件", main, main_data_rows)

    sources: list[tuple[str, TableData]] = []
    source_data_rows: dict[str, list[tuple[int, list[object]]]] = {}
    for role, path in selected_sources:
        logger(f"[读取] 正在读取{role}：{path.name}")
        table = load_table(path, base_fields, metrics)
        rows = _effective_data_rows(table, base_fields, metrics)
        log_table_profile(role, table, rows)
        sources.append((role, table))
        source_data_rows[role] = rows

    workbook, sheet = _create_output_workbook(main)
    try:
        if main.summary_row_index is not None:
            sheet.delete_rows(main.summary_row_index + 1, 1)

        original_width = main_columns
        if original_width < 1:
            raise ValueError("主文件表头没有有效列")
        header_row = main.header_index + 1
        output_columns: dict[tuple[str, str], int] = {}
        next_column = original_width + 1
        for role, _table in sources:
            for metric in metrics:
                output_columns[(role, normalize_header(metric))] = next_column
                _copy_header_style(sheet, original_width, next_column, header_row)
                sheet.cell(header_row, next_column, f"{role}{metric.strip()}")
                next_column += 1

        added_columns = [
            f"{role}{metric}=第{output_columns[(role, normalize_header(metric))]}列"
            for role, _table in sources
            for metric in metrics
        ]
        logger(
            f"[建表] 主表原有 {original_width} 个有效列；新增 {len(added_columns)} 个角色指标列；"
            f"输出共 {next_column - 1} 列"
        )
        logger(f"[建表] 新增列明细：{'；'.join(added_columns)}")

        main_base_columns = {
            normalize_header(field): _column_index(main, field, use_last=False) + 1
            for field in base_fields
        }
        append_row = _last_content_row(main.rows, main.summary_row_index) + 1
        total_rows = sum(len(source_data_rows[role]) for role, _table in sources)
        completed = 0
        counts: dict[str, int] = {}

        for role, table in sources:
            logger(
                f"[追加] 开始处理{role}：准备追加 {len(source_data_rows[role])} 行，"
                f"起始写入输出第 {append_row} 行"
            )
            base_indexes = {normalize_header(field): _column_index(table, field, False) for field in base_fields}
            metric_indexes = {normalize_header(metric): _column_index(table, metric, True) for metric in metrics}
            count = 0
            for source_row_index, row in source_data_rows[role]:
                for field_name, source_index in base_indexes.items():
                    _write_value(
                        sheet,
                        append_row,
                        main_base_columns[field_name],
                        _cell_value(row, source_index),
                        _cell_number_format(table, source_row_index, source_index),
                    )
                for metric_name, source_index in metric_indexes.items():
                    _write_value(
                        sheet,
                        append_row,
                        output_columns[(role, metric_name)],
                        _cell_value(row, source_index),
                        _cell_number_format(table, source_row_index, source_index),
                    )
                append_row += 1
                count += 1
                completed += 1
                if progress:
                    progress(completed, max(total_rows, 1))
            counts[role] = count
            logger(f"[追加] {role}完成：已追加 {count} 行，下一写入行为第 {append_row} 行")

        final_data_rows = len(main_data_rows) + sum(counts.values())
        final_columns = next_column - 1
        final_sheet_rows = append_row - 1
        output.parent.mkdir(parents=True, exist_ok=True)
        logger(f"[保存] 正在写入输出文件：{output.resolve()}")
        workbook.save(output)
        logger(
            f"[完成] 最终结果：有效数据 {final_data_rows} 行 × {final_columns} 列；"
            f"工作表共 {final_sheet_rows} 行（含表头及表头前说明行）"
        )
        logger(f"[完成] 各角色追加行数：{'；'.join(f'{role}={count}行' for role, count in counts.items())}")
        logger(f"[完成] 输出文件：{output.resolve()}")
        logger("=" * 72)
        return MergeResult(output, counts)
    finally:
        workbook.close()


class WeeklyReportWindow:
    """独立周报合并窗口。"""

    def __init__(self, parent: tk.Misc):
        self.window = tk.Toplevel(parent)
        configure_theme(self.window)
        self.window.title("Image Weekly Studio · 周报合并")
        self.window.geometry("1040x900")
        self.window.minsize(900, 760)
        self.messages: queue.Queue[dict[str, object]] = queue.Queue()
        self.file_vars = {"周报本月": tk.StringVar(), **{role: tk.StringVar() for role in SOURCE_ROLES}}
        self._build_ui()
        self._poll_messages()

    def _build_ui(self) -> None:
        container = ttk.Frame(self.window, padding=(20, 14), style="App.TFrame")
        container.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(container, padding=(24, 14), style="Header.TFrame")
        header.pack(fill=tk.X, pady=(0, 10))
        header_copy = ttk.Frame(header, style="Header.TFrame")
        header_copy.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(header_copy, text="周报合并", style="HeaderTitle.TLabel").pack(anchor=tk.W)
        ttk.Label(
            header_copy,
            text="按角色纵向追加数据 · 自动识别表头并保留数值格式",
            style="HeaderSubtitle.TLabel",
        ).pack(anchor=tk.W, pady=(4, 0))
        self.merge_button = ttk.Button(
            header,
            text="开始合并",
            command=self._start_merge,
            style="Header.Primary.TButton",
        )
        self.merge_button.pack(side=tk.RIGHT, padx=(18, 0))

        files_frame = ttk.LabelFrame(
            container,
            text="01  选择文件",
            padding=(14, 8),
            style="Card.TLabelframe",
        )
        files_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(
            files_frame,
            text="周报本月为主文件（必选）；同期月、本周、上周、同期周可按需选择。",
            style="CardText.TLabel",
        ).grid(row=0, column=0, columnspan=4, sticky=tk.W, pady=(0, 4))
        for row, role in enumerate(("周报本月",) + SOURCE_ROLES, start=1):
            label = "主文件 · 周报本月" if role == "周报本月" else role
            ttk.Label(files_frame, text=label, width=18, style="CardText.TLabel").grid(
                row=row, column=0, sticky=tk.W, pady=2
            )
            ttk.Entry(
                files_frame,
                textvariable=self.file_vars[role],
                style="Compact.Modern.TEntry",
            ).grid(row=row, column=1, sticky=tk.EW, padx=(6, 8), pady=2)
            ttk.Button(
                files_frame,
                text="选择文件",
                command=lambda r=role: self._choose_file(r),
                style="Secondary.TButton",
                padding=(10, 4),
            ).grid(row=row, column=2, pady=2)
            ttk.Button(
                files_frame,
                text="清除",
                command=lambda r=role: self.file_vars[r].set(""),
                style="Ghost.TButton",
                padding=(7, 4),
            ).grid(row=row, column=3, padx=(4, 0), pady=2)
        files_frame.columnconfigure(1, weight=1)

        config_frame = ttk.LabelFrame(
            container,
            text="02  配置字段",
            padding=(14, 8),
            style="Card.TLabelframe",
        )
        config_frame.pack(fill=tk.X, expand=False, pady=(0, 10))
        config_frame.columnconfigure(0, weight=1, uniform="fields")
        config_frame.columnconfigure(1, weight=1, uniform="fields")
        config_frame.rowconfigure(1, weight=1)

        ttk.Label(
            config_frame,
            text="每行一个字段，也支持逗号分隔；重复同名指标自动读取最后一列。",
            style="CardText.TLabel",
        ).grid(row=0, column=0, sticky=tk.W, pady=(0, 4))
        ttk.Button(
            config_frame,
            text="恢复默认字段",
            command=self._restore_defaults,
            style="Ghost.TButton",
        ).grid(row=0, column=1, sticky=tk.E, pady=(0, 4))

        base_panel = ttk.Frame(config_frame, style="Card.TFrame")
        base_panel.grid(row=1, column=0, sticky=tk.NSEW, padx=(0, 6))
        base_panel.columnconfigure(0, weight=1)
        base_panel.rowconfigure(1, weight=1)
        ttk.Label(base_panel, text="基础字段 · 默认 14 个", style="CardTitle.TLabel").grid(
            row=0, column=0, sticky=tk.W, pady=(0, 5)
        )
        self.base_text = tk.Text(base_panel, height=4, wrap=tk.NONE)
        style_text_widget(self.base_text)
        self.base_text.grid(row=1, column=0, sticky=tk.NSEW)
        base_scrollbar = ttk.Scrollbar(
            base_panel,
            orient=tk.VERTICAL,
            command=self.base_text.yview,
            style="Modern.Vertical.TScrollbar",
        )
        base_scrollbar.grid(row=1, column=1, sticky=tk.NS)
        self.base_text.configure(yscrollcommand=base_scrollbar.set)

        metric_panel = ttk.Frame(config_frame, style="Card.TFrame")
        metric_panel.grid(row=1, column=1, sticky=tk.NSEW, padx=(6, 0))
        metric_panel.columnconfigure(0, weight=1)
        metric_panel.rowconfigure(1, weight=1)
        ttk.Label(metric_panel, text="追加指标 · 默认 5 个", style="CardTitle.TLabel").grid(
            row=0, column=0, sticky=tk.W, pady=(0, 5)
        )
        self.metric_text = tk.Text(metric_panel, height=4, wrap=tk.NONE)
        style_text_widget(self.metric_text)
        self.metric_text.grid(row=1, column=0, sticky=tk.NSEW)
        metric_scrollbar = ttk.Scrollbar(
            metric_panel,
            orient=tk.VERTICAL,
            command=self.metric_text.yview,
            style="Modern.Vertical.TScrollbar",
        )
        metric_scrollbar.grid(row=1, column=1, sticky=tk.NS)
        self.metric_text.configure(yscrollcommand=metric_scrollbar.set)
        self._restore_defaults()

        status_frame = ttk.Frame(container, padding=(16, 11), style="Card.TFrame")
        status_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(status_frame, text="合并进度", style="CardTitle.TLabel").pack(
            side=tk.LEFT, padx=(0, 12)
        )
        self.progress = ttk.Progressbar(
            status_frame,
            mode="determinate",
            style="Business.Horizontal.TProgressbar",
        )
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 12))
        self.progress_label = ttk.Label(status_frame, text="就绪", style="Status.TLabel")
        self.progress_label.pack(side=tk.RIGHT)

        log_frame = ttk.LabelFrame(
            container,
            text="03  处理日志",
            padding=10,
            style="Card.TLabelframe",
        )
        log_frame.pack(fill=tk.BOTH, expand=True)
        log_toolbar = ttk.Frame(log_frame, style="Card.TFrame")
        log_toolbar.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(
            log_toolbar,
            text="完整记录文件、字段匹配、汇总行、追加进度和最终行列数",
            style="CardText.TLabel",
        ).pack(side=tk.LEFT)
        ttk.Button(
            log_toolbar,
            text="复制日志",
            command=self._copy_log,
            style="Secondary.TButton",
            padding=(10, 4),
        ).pack(side=tk.RIGHT)
        ttk.Button(
            log_toolbar,
            text="清空",
            command=self._clear_log,
            style="Ghost.TButton",
            padding=(7, 4),
        ).pack(side=tk.RIGHT, padx=(0, 6))

        log_body = ttk.Frame(log_frame, style="Card.TFrame")
        log_body.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(log_body, height=14, wrap=tk.WORD, state=tk.DISABLED)
        style_text_widget(self.log_text)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(
            log_body,
            orient=tk.VERTICAL,
            command=self.log_text.yview,
            style="Modern.Vertical.TScrollbar",
        )
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        footer = ttk.Frame(container, style="App.TFrame")
        footer.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(
            footer,
            text="输出始终保存为新的 XLSX 文件，不会覆盖输入文件。",
            style="Muted.TLabel",
        ).pack(side=tk.LEFT)
        ttk.Button(
            footer,
            text="关闭窗口",
            command=self.window.destroy,
            style="Ghost.TButton",
        ).pack(side=tk.RIGHT)

    def _choose_file(self, role: str) -> None:
        path = filedialog.askopenfilename(
            parent=self.window,
            title=f"选择{role}文件",
            filetypes=[("周报文件", "*.xlsx *.xls *.csv"), ("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")],
        )
        if path:
            self.file_vars[role].set(path)

    def _restore_defaults(self) -> None:
        self.base_text.delete("1.0", tk.END)
        self.base_text.insert("1.0", "\n".join(DEFAULT_BASE_FIELDS))
        self.metric_text.delete("1.0", tk.END)
        self.metric_text.insert("1.0", "\n".join(DEFAULT_METRICS))

    def _start_merge(self) -> None:
        main_path = self.file_vars["周报本月"].get().strip()
        sources = {role: self.file_vars[role].get().strip() for role in SOURCE_ROLES}
        base_fields = parse_field_text(self.base_text.get("1.0", tk.END))
        metrics = parse_field_text(self.metric_text.get("1.0", tk.END))
        if not main_path:
            messagebox.showwarning("缺少主文件", "请选择主文件（周报本月）。", parent=self.window)
            return
        if not any(sources.values()):
            messagebox.showwarning("缺少追加文件", "请至少选择一个追加文件。", parent=self.window)
            return
        if not base_fields or not metrics:
            messagebox.showwarning("字段为空", "基础字段和追加指标都不能为空。", parent=self.window)
            return

        output_path = filedialog.asksaveasfilename(
            parent=self.window,
            title="保存周报合并结果",
            initialdir=str(Path(main_path).parent),
            initialfile="周报合并结果.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if not output_path:
            return
        self.merge_button.configure(state=tk.DISABLED)
        self.progress.configure(value=0, maximum=1)
        self.progress_label.configure(text="准备中")
        self._append_log("开始合并……")
        worker = threading.Thread(
            target=self._merge_worker,
            args=(main_path, sources, base_fields, metrics, output_path),
            daemon=True,
        )
        worker.start()

    def _merge_worker(
        self,
        main_path: str,
        sources: dict[str, str],
        base_fields: list[str],
        metrics: list[str],
        output_path: str,
    ) -> None:
        try:
            result = merge_reports(
                main_path,
                sources,
                base_fields,
                metrics,
                output_path,
                log=lambda text: self.messages.put({"type": "log", "text": text}),
                progress=lambda value, maximum: self.messages.put(
                    {"type": "progress", "value": value, "maximum": maximum}
                ),
            )
            self.messages.put({"type": "done", "result": result})
        except Exception as error:
            self.messages.put({"type": "error", "text": str(error)})

    def _poll_messages(self) -> None:
        if not self.window.winfo_exists():
            return
        try:
            while True:
                message = self.messages.get_nowait()
                message_type = message["type"]
                if message_type == "log":
                    self._append_log(str(message["text"]))
                elif message_type == "progress":
                    maximum = int(message["maximum"])
                    value = int(message["value"])
                    self.progress.configure(maximum=maximum, value=value)
                    self.progress_label.configure(text=f"{value}/{maximum}")
                elif message_type == "done":
                    self._handle_done(message["result"])
                elif message_type == "error":
                    self._handle_error(str(message["text"]))
        except queue.Empty:
            pass
        self.window.after(100, self._poll_messages)

    def _append_log(self, text: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        lines = text.splitlines() or [""]
        self.log_text.configure(state=tk.NORMAL)
        for line in lines:
            self.log_text.insert(tk.END, f"[{timestamp}] {line}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _clear_log(self) -> None:
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _copy_log(self) -> None:
        content = self.log_text.get("1.0", tk.END).strip()
        if not content:
            return
        self.window.clipboard_clear()
        self.window.clipboard_append(content)
        self.window.update_idletasks()
        self._append_log("日志已复制到剪贴板")

    def _handle_done(self, result: object) -> None:
        assert isinstance(result, MergeResult)
        self.merge_button.configure(state=tk.NORMAL)
        counts = "\n".join(f"{role}：{count} 行" for role, count in result.source_counts.items())
        self.progress_label.configure(text="完成")
        messagebox.showinfo(
            "合并完成",
            f"周报合并成功。\n\n{counts}\n\n输出文件：\n{result.output_path}",
            parent=self.window,
        )

    def _handle_error(self, text: str) -> None:
        self.merge_button.configure(state=tk.NORMAL)
        self.progress_label.configure(text="失败")
        self._append_log(f"错误：{text}")
        messagebox.showerror("合并失败", text, parent=self.window)
